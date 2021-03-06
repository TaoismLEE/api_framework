# _*_ coding:utf-8 _*_

import os
import shutil
from config import setting
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
import configparser as cparser

# --------- 读取config.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read(setting.TEST_CONFIG, encoding='UTF-8')
name = cf.get("tester", "name")


class WriteExcel:
    """文件写入数据"""
    def __init__(self, report_file, case_file):
        self.filename = report_file
        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝模板文件至指定报告目录下
            shutil.copyfile(case_file, report_file)
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write_data(self, row_n, value, response):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :para response: 接口返回结果
        :return: 无
        """
        font_GREEN = Font(name='宋体', color="00228B22", bold=True)
        font_RED = Font(name='宋体', color="00FF0000", bold=True)
        font1 = Font(name='宋体', color="00A52A2A", bold=True)
        align = Alignment(horizontal='center', vertical='center')
        # 获数所在行数
        L_n = "L" + str(row_n)
        M_n = "M" + str(row_n)
        if value == "PASS":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_GREEN
        if value == "FAIL":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_RED
        self.ws.cell(row_n, 13, name)
        self.ws.cell(row_n, 14, str(response))
        self.ws[L_n].alignment = align
        self.ws[M_n].font = font1
        self.ws[M_n].alignment = align
        self.wb.save(self.filename)
